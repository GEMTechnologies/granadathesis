"""
Spreadsheet Service - Excel and CSV Read/Write with Formulas and Charts

Features:
- Read Excel (.xlsx, .xls) and CSV files to DataFrame
- Write data with formulas (SUM, AVERAGE, COUNT, etc.)
- Auto-create charts (bar, line, pie, scatter)
- Export charts as images

Dependencies: pandas, openpyxl, xlsxwriter, matplotlib
"""

import os
import io
from pathlib import Path
from typing import Dict, Any, List, Optional, Union, Tuple
import json

# Data handling
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

# Excel writing with formulas
try:
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

# Charts and graphs
try:
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    plt = None


class SpreadsheetService:
    """Service for reading, writing, and analyzing spreadsheet data."""
    
    def __init__(self, workspace_dir: str = "workspaces"):
        self.workspace_dir = Path(workspace_dir)
        self.workspace_dir.mkdir(parents=True, exist_ok=True)
    
    # =========================================================================
    # READ OPERATIONS
    # =========================================================================
    
    def read_excel(
        self, 
        file_path: Union[str, Path], 
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Read Excel file to DataFrame(s).
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet to read (None = all sheets)
            
        Returns:
            {
                "success": True,
                "sheets": {"Sheet1": [...data...], "Sheet2": [...data...]},
                "columns": {"Sheet1": [...], "Sheet2": [...]},
                "row_count": {"Sheet1": 100, "Sheet2": 50}
            }
        """
        if not PANDAS_AVAILABLE:
            return {"success": False, "error": "pandas not installed"}
        
        try:
            file_path = Path(file_path)
            
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                return {
                    "success": True,
                    "sheets": {sheet_name: df.to_dict('records')},
                    "columns": {sheet_name: list(df.columns)},
                    "row_count": {sheet_name: len(df)},
                    "dataframe": df  # For internal use
                }
            else:
                # Read all sheets
                excel_file = pd.ExcelFile(file_path)
                sheets_data = {}
                columns = {}
                row_counts = {}
                dataframes = {}
                
                for sheet in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet)
                    sheets_data[sheet] = df.to_dict('records')
                    columns[sheet] = list(df.columns)
                    row_counts[sheet] = len(df)
                    dataframes[sheet] = df
                
                return {
                    "success": True,
                    "sheets": sheets_data,
                    "columns": columns,
                    "row_count": row_counts,
                    "dataframes": dataframes
                }
                
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def read_csv(
        self, 
        file_path: Union[str, Path],
        delimiter: str = ","
    ) -> Dict[str, Any]:
        """
        Read CSV file to DataFrame.
        
        Args:
            file_path: Path to CSV file
            delimiter: Column delimiter
            
        Returns:
            {
                "success": True,
                "data": [...records...],
                "columns": [...],
                "row_count": 100
            }
        """
        if not PANDAS_AVAILABLE:
            return {"success": False, "error": "pandas not installed"}
        
        try:
            df = pd.read_csv(file_path, delimiter=delimiter)
            return {
                "success": True,
                "data": df.to_dict('records'),
                "columns": list(df.columns),
                "row_count": len(df),
                "dataframe": df
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    # =========================================================================
    # WRITE OPERATIONS
    # =========================================================================
    
    def write_excel(
        self,
        data: Union[List[Dict], 'pd.DataFrame'],
        file_path: Union[str, Path],
        sheet_name: str = "Sheet1",
        formulas: Optional[Dict[str, str]] = None,
        auto_formulas: Optional[List[str]] = None,
        create_chart: Optional[str] = None,
        workspace_id: str = "default"
    ) -> Dict[str, Any]:
        """
        Write data to Excel with optional formulas and charts.
        
        Args:
            data: List of dicts or DataFrame
            file_path: Output file path (relative to workspace)
            sheet_name: Sheet name
            formulas: {"B10": "=SUM(B2:B9)", "C10": "=AVERAGE(C2:C9)"}
            auto_formulas: ["SUM", "AVERAGE"] - auto-apply to numeric columns
            create_chart: "bar", "line", "pie", "scatter"
            
        Returns:
            {"success": True, "path": "...", "chart_path": "..."}
        """
        if not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "pandas/openpyxl not installed"}
        
        try:
            # Convert to DataFrame if needed
            if isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = data
            
            # Determine output path
            workspace_path = self.workspace_dir / workspace_id
            workspace_path.mkdir(parents=True, exist_ok=True)
            
            if not str(file_path).endswith('.xlsx'):
                file_path = str(file_path) + '.xlsx'
            
            output_path = workspace_path / file_path
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Write headers
            for col_idx, column in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=column)
            
            # Write data
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    # Convert numpy types to Python types
                    if hasattr(value, 'item'):
                        value = value.item()
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply manual formulas
            if formulas:
                for cell_ref, formula in formulas.items():
                    ws[cell_ref] = formula
            
            # Apply auto-formulas
            if auto_formulas:
                last_row = len(df) + 1
                formula_row = last_row + 1
                
                for col_idx, column in enumerate(df.columns, 1):
                    # Only apply to numeric columns
                    if df[column].dtype in ['int64', 'float64']:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        
                        if "SUM" in auto_formulas:
                            ws.cell(row=formula_row, column=col_idx, 
                                   value=f"=SUM({col_letter}2:{col_letter}{last_row})")
                        if "AVERAGE" in auto_formulas:
                            ws.cell(row=formula_row + 1, column=col_idx,
                                   value=f"=AVERAGE({col_letter}2:{col_letter}{last_row})")
                        if "COUNT" in auto_formulas:
                            ws.cell(row=formula_row + 2, column=col_idx,
                                   value=f"=COUNT({col_letter}2:{col_letter}{last_row})")
                        if "MAX" in auto_formulas:
                            ws.cell(row=formula_row + 3, column=col_idx,
                                   value=f"=MAX({col_letter}2:{col_letter}{last_row})")
                        if "MIN" in auto_formulas:
                            ws.cell(row=formula_row + 4, column=col_idx,
                                   value=f"=MIN({col_letter}2:{col_letter}{last_row})")
            
            # Create embedded chart if requested
            chart_path = None
            if create_chart and MATPLOTLIB_AVAILABLE:
                chart_path = self._create_chart_in_excel(ws, df, create_chart, len(df))
            
            # Save workbook
            wb.save(output_path)
            
            result = {
                "success": True,
                "path": str(output_path),
                "relative_path": f"{workspace_id}/{file_path}",
                "rows": len(df),
                "columns": len(df.columns)
            }
            
            if chart_path:
                result["chart_path"] = chart_path
            
            return result
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {"success": False, "error": str(e)}
    
    def write_csv(
        self,
        data: Union[List[Dict], 'pd.DataFrame'],
        file_path: Union[str, Path],
        workspace_id: str = "default",
        delimiter: str = ","
    ) -> Dict[str, Any]:
        """
        Write data to CSV file.
        
        Args:
            data: List of dicts or DataFrame
            file_path: Output file path
            workspace_id: Workspace ID
            delimiter: Column delimiter
            
        Returns:
            {"success": True, "path": "..."}
        """
        if not PANDAS_AVAILABLE:
            return {"success": False, "error": "pandas not installed"}
        
        try:
            if isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = data
            
            workspace_path = self.workspace_dir / workspace_id
            workspace_path.mkdir(parents=True, exist_ok=True)
            
            if not str(file_path).endswith('.csv'):
                file_path = str(file_path) + '.csv'
            
            output_path = workspace_path / file_path
            df.to_csv(output_path, index=False, sep=delimiter)
            
            return {
                "success": True,
                "path": str(output_path),
                "relative_path": f"{workspace_id}/{file_path}",
                "rows": len(df),
                "columns": len(df.columns)
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    # =========================================================================
    # CHART/GRAPH OPERATIONS
    # =========================================================================
    
    def _create_chart_in_excel(
        self, 
        ws, 
        df: 'pd.DataFrame', 
        chart_type: str,
        data_rows: int
    ) -> Optional[str]:
        """Create chart embedded in Excel worksheet."""
        try:
            # Determine numeric columns for chart
            numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns.tolist()
            if not numeric_cols:
                return None
            
            # Create chart based on type
            if chart_type == "bar":
                chart = BarChart()
                chart.type = "col"
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            elif chart_type == "scatter":
                chart = ScatterChart()
            else:
                chart = BarChart()
            
            chart.title = "Data Chart"
            
            # Add data references
            data_col = df.columns.get_loc(numeric_cols[0]) + 1
            data = Reference(ws, min_col=data_col, min_row=1, max_row=data_rows + 1)
            categories = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Position chart
            ws.add_chart(chart, f"{openpyxl.utils.get_column_letter(len(df.columns) + 2)}2")
            
            return "embedded"
            
        except Exception as e:
            print(f"Chart error: {e}")
            return None
    
    def create_chart(
        self,
        data: Union[List[Dict], 'pd.DataFrame', Dict],
        chart_type: str,
        title: str = "Chart",
        x_column: Optional[str] = None,
        y_columns: Optional[List[str]] = None,
        workspace_id: str = "default",
        filename: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Create a standalone chart image.
        
        Args:
            data: Data for chart
            chart_type: "bar", "line", "pie", "scatter", "histogram"
            title: Chart title
            x_column: Column for x-axis
            y_columns: Columns for y-axis
            workspace_id: Workspace ID
            filename: Output filename (auto-generated if None)
            
        Returns:
            {"success": True, "path": "...", "relative_path": "..."}
        """
        if not MATPLOTLIB_AVAILABLE:
            return {"success": False, "error": "matplotlib not installed"}
        
        try:
            # Convert to DataFrame
            if isinstance(data, list):
                df = pd.DataFrame(data)
            elif isinstance(data, dict):
                df = pd.DataFrame(data)
            else:
                df = data
            
            # Auto-detect columns if not specified
            if not x_column and len(df.columns) > 0:
                x_column = df.columns[0]
            
            if not y_columns:
                y_columns = df.select_dtypes(include=['int64', 'float64']).columns.tolist()
            
            # Create figure
            fig, ax = plt.subplots(figsize=(10, 6))
            
            if chart_type == "bar":
                df.plot(kind='bar', x=x_column, y=y_columns, ax=ax)
            elif chart_type == "line":
                df.plot(kind='line', x=x_column, y=y_columns, ax=ax)
            elif chart_type == "pie":
                if y_columns:
                    df.set_index(x_column)[y_columns[0]].plot(kind='pie', ax=ax, autopct='%1.1f%%')
            elif chart_type == "scatter":
                if len(y_columns) >= 2:
                    ax.scatter(df[y_columns[0]], df[y_columns[1]])
                    ax.set_xlabel(y_columns[0])
                    ax.set_ylabel(y_columns[1])
            elif chart_type == "histogram":
                if y_columns:
                    df[y_columns[0]].plot(kind='hist', ax=ax, bins=20)
            else:
                # Auto-detect best chart
                df.plot(ax=ax)
            
            ax.set_title(title)
            plt.tight_layout()
            
            # Save chart
            workspace_path = self.workspace_dir / workspace_id / "charts"
            workspace_path.mkdir(parents=True, exist_ok=True)
            
            if not filename:
                import time
                filename = f"chart_{int(time.time())}.png"
            elif not filename.endswith('.png'):
                filename += '.png'
            
            output_path = workspace_path / filename
            plt.savefig(output_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            return {
                "success": True,
                "path": str(output_path),
                "relative_path": f"{workspace_id}/charts/{filename}",
                "chart_type": chart_type
            }
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {"success": False, "error": str(e)}
    
    def auto_analyze(
        self,
        data: Union[List[Dict], 'pd.DataFrame'],
        workspace_id: str = "default"
    ) -> Dict[str, Any]:
        """
        Automatically analyze data and generate appropriate charts.
        
        Returns:
            {
                "success": True,
                "summary": {...statistics...},
                "charts": [...chart paths...],
                "insights": [...AI insights...]
            }
        """
        if not PANDAS_AVAILABLE:
            return {"success": False, "error": "pandas not installed"}
        
        try:
            if isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = data
            
            # Generate summary statistics
            summary = {
                "shape": {"rows": len(df), "columns": len(df.columns)},
                "columns": list(df.columns),
                "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
                "numeric_summary": df.describe().to_dict() if len(df.select_dtypes(include=['number']).columns) > 0 else {}
            }
            
            charts = []
            
            # Auto-generate charts based on data
            numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns.tolist()
            
            if len(numeric_cols) >= 1:
                # Bar chart for first numeric column
                result = self.create_chart(
                    df, "bar", f"{numeric_cols[0]} Distribution",
                    x_column=df.columns[0], y_columns=[numeric_cols[0]],
                    workspace_id=workspace_id, filename=f"{numeric_cols[0]}_bar.png"
                )
                if result.get("success"):
                    charts.append(result["relative_path"])
            
            if len(numeric_cols) >= 2:
                # Scatter plot for correlation
                result = self.create_chart(
                    df, "scatter", f"{numeric_cols[0]} vs {numeric_cols[1]}",
                    y_columns=numeric_cols[:2],
                    workspace_id=workspace_id, filename="correlation_scatter.png"
                )
                if result.get("success"):
                    charts.append(result["relative_path"])
            
            return {
                "success": True,
                "summary": summary,
                "charts": charts,
                "insights": [
                    f"Dataset has {summary['shape']['rows']} rows and {summary['shape']['columns']} columns",
                    f"Numeric columns: {', '.join(numeric_cols)}" if numeric_cols else "No numeric columns found"
                ]
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}


# Singleton instance
_spreadsheet_service = None

def get_spreadsheet_service() -> SpreadsheetService:
    """Get global spreadsheet service instance."""
    global _spreadsheet_service
    if _spreadsheet_service is None:
        _spreadsheet_service = SpreadsheetService()
    return _spreadsheet_service
